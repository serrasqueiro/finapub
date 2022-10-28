#-*- coding: utf-8 -*-
# euronextimport.py  (c)2020  Henrique Moreira

""" Imports Euronext list of stocks
"""

import sys
import openpyxl
from waxpage.redit import char_map
from mintracker.sindexes.isin import ISIN

# pylint: disable=missing-function-docstring, line-too-long, use-list-literal

DEF_EN_EQ_FNAME = "Euronext_Equities.xlsx"
LINEAR_DUMP = False	# use True if you want to generate EURONEXT_STOCKS linearly

MKT_MAP = {
    "BRU.A": "Euronext Access Brussels",
    "LIS.A": "Euronext Access Lisbon",
    "PAR.A": "Euronext Access Paris",
    "AMS.A": "Euronext Amsterdam",
    "EN.AB": "Euronext Amsterdam, Brussels",
    "EN.ABP": "Euronext Amsterdam, Brussels, Paris",
    "EN.AP": "Euronext Amsterdam, Paris",
    "EN.B": "Euronext Brussels",
    "EN.BA": "Euronext Brussels, Amsterdam",
    "EN.BP": "Euronext Brussels, Paris",
    "EN.DUBLIN": "Euronext Dublin",
    "EX.OSLO": "Euronext Expand Oslo",
    "EN.EXP": "Euronext Expert Market",
    "EG.BRU": "Euronext Growth Brussels",
    "EG.DUBLIN": "Euronext Growth Dublin",
    "EG.LIS": "Euronext Growth Lisbon",
    "EG.OSLO": "Euronext Growth Oslo",
    "EG.PAR": "Euronext Growth Paris",
    "EG.PB": "Euronext Growth Paris, Brussels",
    "EN.LIS": "Euronext Lisbon",
    "EN.PAR": "Euronext Paris",
    "EN.PA": "Euronext Paris, Amsterdam",
    "EN.PAB": "Euronext Paris, Amsterdam, Brussels",
    "EN.PB": "Euronext Paris, Brussels",
    "OSLO": "Oslo B.rs",
    "NLB": "Traded not listed Brussels",
}

# date: 24.12.2020
# source https://live.euronext.com/pt/products/equities/list
#	Euronext_Equities_YYYY-MM-DD.xls
#
# Formats: xls, csv, or txt
# Headers (columns):
#	1. Name
#	2. ISIN
#	3. Symbol
#	4. Market
#		Euronext (...), or Oslo Bors
#	5. Trading Currency
#	6. Open
#	7. High
#	8. Low
#	9. Last
#	10. Last Date/Time
#	11. Time Zone
#	12. Volume
#	13. Turnover
#
# CSV header:
#	'Name;ISIN;Symbol;Market;"Trading Currency";Open;High;Low;Last;"Last Date/Time";"Time Zone";Volume;Turnover'


def main():
    """ Main script """
    myprog = __file__
    args = sys.argv[1:]
    code = runner(args)
    if code is None:
        print(f"""Usage:
{myprog} Euronext_Equities_...xlsx [hints]

Hints are:
   EUR - only display Euro stocks
""")
    sys.exit(code if code else 0)


def runner (args, debug=0):
    """ Run importer """
    param = args
    filtered = None
    if not args:
        return None
    fname = param[0]
    del param[0]
    if param:
        filtered = param[0].upper()
        del param[0]
    assert not param
    imp = Importer(fname)
    msgs = imp.get_messages()
    shown = '\n'.join(msgs)
    if msgs:
        print("Note!\n")
        print(f"{shown}")
        return 1
    opts = {"filter": filtered,
            "pre": " " * 4,
            }
    if LINEAR_DUMP:
        linear_dump(imp, opts, sys.stdout, debug)
        return 0
    dump_by_index(imp, opts, sys.stdout, debug)
    return 0


def linear_dump(imp, opts, out, debug=0):
    dump_import(imp, opts, out)
    num = len(imp.content)
    if debug > 0:
        print(f"Dumped #{num} equities, see EURONEXT_STOCKS tuple at euronext.py")


def dump_by_index(imp, opts, out, debug=0):
    """ Dump by market index """
    dct = dump_import(imp, opts, None)
    markets = sorted(dct["markets"])
    m_names = [item for _, item in MKT_MAP.items()]
    if debug > 0:
        print(f"Debug: markets={markets}\n")
        for m_name in m_names:
            mkt = short_market_name(m_name)
            isin_list = dct["market-isin"].get(mkt)
            if isin_list is None:
                print(f"Debug: note '{mkt}' ('{m_name}') is not mapped")
                continue
            assert isinstance(isin_list, (list, tuple))
            shown_isin = isin_list if len(isin_list) < 3 else (isin_list[:4] + ["..."])
            print(f"market-isin[{mkt}] len#{len(isin_list)}:", shown_isin)
    avar = {}
    mkts = sorted(MKT_MAP.keys())
    shown = ""
    for mkt in mkts:
        shown += (" " * 4) + f"'{mkt}',\n"
    out.write(f"EURONEXT_STOCKS_LIST = (\n{shown}    )\n\n")
    for mkt in mkts:
        abbrev = market_to_varname(mkt)
        varname = f"EURONEXT_STOCKS_{abbrev}"
        avar[varname] = list()
        isin_list = dct["market-isin"].get(mkt)
        if isin_list is None:
            continue
        assert isin_list
        for item in isin_list:
            count, tup = 0, None
            for stock in dct["list"]:
                isin = stock[1]
                if isin == item:
                    tup = stock
                    count += 1
            #print("Debug: market", mkt, "item:", item, "count:", count, "tup:", tup)
            assert count > 0
            if count > 1:
                continue
            #print(f"mkt={mkt}: {tup}")
            avar[varname].append(tup)
    for varname in sorted(avar.keys()):
        cont = avar[varname]
        if not cont:
            continue
        out.write(f"{varname} = (\n")
        for tup in cont:
            shown = tup[:-1]
            out.write(f"    {shown},\n")
        out.write("    )\n\n")
        abbrevs, dups = {}, []
        for tup in cont:
            fourplet = tup[:-1]
            coin, isin, symbol, name = fourplet
            assert name
            assert coin
            myisin = ISIN(isin)
            assert myisin.is_valid()
            assert symbol
            if symbol == "-":
                continue
            # if symbol in abbrevs or symbol == "JMT" ...
            if symbol in abbrevs:
                dups.append(f"'{symbol}'")
            else:
                abbrevs[symbol] = fourplet
        if dups:
            shown = ", ".join(dups)
            out.write(f"DUPS_{varname} = ({shown},)\n\n")


def dump_import(imp, opts, out=None) -> dict:
    err = sys.stderr
    stocks = list()
    isins, symbs = {}, {}
    filtered = opts["filter"]
    pre = opts["pre"]
    if pre:
        post = ","
    for row in imp.content:
        alist = char_map.simpler_ascii(row)
        tup = alist[4], alist[1], alist[2], alist[0], alist[3]
        coin, isin, symb = tup[0], tup[1], tup[2]
        if isin in isins:
            err.write(f"Duplicate ISIN {isin}: {isins[isin]}\n")
            continue
        shown = tup[:-1]
        if filtered is None or filtered == coin:
            if out:
                out.write(f"{pre}{shown}{post}\n")
            stocks.append(tup)
        isins[isin] = tup
        if symb == "-":
            continue
        if symb in symbs:
            err.write(f"Duplicate symbol '{symb}', ISIN {isin}: {symbs[symb]}\n")
        symbs[symb] = tup
    res = {
        "list": stocks,
        "markets": [],
        "market-isin": {},
    }
    for stock in stocks:
        market = stock[-1]
        symb = stock[1]
        if market not in res["markets"]:
            res["markets"].append(market)
            mname = short_market_name(market)
            res["market-isin"][mname] = list()
    for stock in stocks:
        mname = short_market_name(stock[-1])
        isin = stock[1]
        res["market-isin"][mname].append(isin)
    return res


class Importer():
    """ XLS importer """
    def __init__(self, fname=""):
        self._msgs = list()
        aname = DEF_EN_EQ_FNAME if not fname else fname
        self.content = self._read(aname)

    def get_messages(self):
        return self._msgs

    def _read(self,fname) -> list:
        wbk = openpyxl.open(fname, read_only=True, data_only=True)
        sheet_name = wbk.sheetnames[0]
        if len(wbk.sheetnames) > 1:
            msg = f"Multiple sheets: {';'.join(wbk.sheetnames)}"
            self._msgs.append(msg)
        sheet = wbk[sheet_name]
        return self._parse_sheet(sheet)

    def _parse_sheet(self, sheet) -> list:
        """ Returns the list of stocks tuples """
        res = list()
        state = -1
        for row in sheet:
            if len(row) < 4:
                continue
            first = row[0].value
            if state == -1:
                if first == "Name":
                    state = 0
                continue
            assert first
            alist = [str(entry.value) for entry in row]
            res.append(alist[:5])
        return res


def short_market_name(market, not_found="?"):
    """ Returns the short name of a market """
    assert isinstance(market, str)
    for key, fullname in MKT_MAP.items():
        if fullname == market:
            return key
    return not_found

def market_to_varname(mkt):
    """ Example... 'EURONEXT_STOCKS_EN_LIS', mkt='EN.LIS'
    Here is only the suffix returned, see also varname_to_market()
    """
    return mkt.replace(".", "_")

def varname_to_market(name):
    return name.replace("_", ".")


# Main script
if __name__ == "__main__":
    main()
